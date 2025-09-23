#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Telegram бот (aiogram 3.x, one-file) для создания Excel с товарами + архив + система одноразовых приглашений.

Запуск:
  pip install aiogram==3.* openpyxl pillow python-dotenv
  export BOT_TOKEN="123:ABC"
  export ADMIN_IDS="123456789,987654321"  # через запятую
  python excel_bot.py

В Replit добавьте переменные среды BOT_TOKEN и ADMIN_IDS в Secrets.

Автор: ChatGPT (один файл, без внешних модулей)
"""
import asyncio
import os
import io
import sqlite3
import textwrap
from datetime import datetime
from dataclasses import dataclass, asdict
from typing import Optional, List, Dict, Any

from aiogram import Bot, Dispatcher, F, Router
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, ContentType, InputMediaPhoto
)
from aiogram.filters import Command, CommandObject
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import InlineKeyboardBuilder

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from PIL import Image as PILImage, UnidentifiedImageError

# -------------------- Конфигурация --------------------
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
ADMIN_IDS = {int(x) for x in os.getenv("ADMIN_IDS", "").replace(" ", "").split(",") if x.isdigit()}

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
ARCHIVE_DIR = os.path.join(BASE_DIR, "archive")
DB_PATH = os.path.join(BASE_DIR, "bot.db")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)

# -------------------- БД --------------------
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS users(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          tg_id INTEGER UNIQUE NOT NULL,
          joined_at TEXT NOT NULL
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS invites(
          code TEXT PRIMARY KEY,
          created_by INTEGER NOT NULL,
          created_at TEXT NOT NULL,
          used_by INTEGER,
          used_at TEXT
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS items(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          user_id INTEGER NOT NULL,
          idx INTEGER NOT NULL,                 -- порядковый номер внутри черновика
          photo_path TEXT,
          link TEXT,
          color TEXT,
          size TEXT,
          qty TEXT,
          comment TEXT,
          status TEXT NOT NULL DEFAULT 'draft'  -- draft/saved (после сохранения Excel можно обнулять)
        );
        """)
        conn.commit()

# -------------------- Утилиты --------------------
def ensure_user_folder(user_id: int) -> str:
    path = os.path.join(DATA_DIR, str(user_id))
    os.makedirs(path, exist_ok=True)
    return path

def ensure_archive_folder(user_id: int) -> str:
    path = os.path.join(ARCHIVE_DIR, str(user_id))
    os.makedirs(path, exist_ok=True)
    return path

def is_registered(tg_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT 1 FROM users WHERE tg_id=?", (tg_id,)).fetchone()
        return row is not None

def register_user(tg_id: int):
    with db() as conn:
        conn.execute("INSERT OR IGNORE INTO users(tg_id, joined_at) VALUES(?, ?)", (tg_id, datetime.utcnow().isoformat()))
        conn.commit()

def use_invite(code: str, tg_id: int) -> bool:
    with db() as conn:
        row = conn.execute("SELECT * FROM invites WHERE code=?", (code,)).fetchone()
        if not row:
            return False
        if row["used_by"] is not None:
            return False
        # помечаем использованным
        conn.execute("UPDATE invites SET used_by=?, used_at=? WHERE code=?", (tg_id, datetime.utcnow().isoformat(), code))
        conn.commit()
        return True

def create_invite(created_by: int) -> str:
    code = f"INV{created_by}-{int(datetime.utcnow().timestamp())}"
    with db() as conn:
        conn.execute("INSERT INTO invites(code, created_by, created_at) VALUES(?, ?, ?)", (code, created_by, datetime.utcnow().isoformat()))
        conn.commit()
    return code

def list_open_invites() -> List[sqlite3.Row]:
    with db() as conn:
        return conn.execute("SELECT * FROM invites WHERE used_by IS NULL ORDER BY created_at DESC").fetchall()

def next_item_index(user_id: int) -> int:
    with db() as conn:
        row = conn.execute("SELECT COALESCE(MAX(idx),0)+1 AS n FROM items WHERE user_id=? AND status='draft'", (user_id,)).fetchone()
        return int(row["n"] or 1)

def read_draft_items(user_id: int) -> List[sqlite3.Row]:
    with db() as conn:
        return conn.execute("""
            SELECT * FROM items WHERE user_id=? AND status='draft' ORDER BY idx
        """, (user_id,)).fetchall()

def get_item_by_idx(user_id: int, idx: int) -> Optional[sqlite3.Row]:
    with db() as conn:
        return conn.execute("""SELECT * FROM items WHERE user_id=? AND status='draft' AND idx=?""", (user_id, idx)).fetchone()

def upsert_item_field(user_id: int, idx: int, field: str, value: Optional[str]):
    with db() as conn:
        row = conn.execute("SELECT id FROM items WHERE user_id=? AND status='draft' AND idx=?", (user_id, idx)).fetchone()
        if row is None:
            conn.execute("""INSERT INTO items(user_id, idx, {f}) VALUES(?,?,?)""".format(f=field), (user_id, idx, value))
        else:
            conn.execute(f"UPDATE items SET {field}=? WHERE user_id=? AND idx=? AND status='draft'", (value, user_id, idx))
        conn.commit()

def delete_item(user_id: int, idx: int):
    with db() as conn:
        # удаляем физически фото, если есть
        row = conn.execute("SELECT photo_path FROM items WHERE user_id=? AND idx=? AND status='draft'", (user_id, idx)).fetchone()
        if row and row["photo_path"] and os.path.exists(row["photo_path"]):
            try: os.remove(row["photo_path"])
            except: pass
        conn.execute("DELETE FROM items WHERE user_id=? AND idx=? AND status='draft'", (user_id, idx))
        # сдвигаем индексы выше
        conn.execute("""
            UPDATE items SET idx = idx - 1
            WHERE user_id=? AND status='draft' AND idx > ?
        """, (user_id, idx))
        conn.commit()

def reset_draft(user_id: int):
    with db() as conn:
        rows = conn.execute("SELECT photo_path FROM items WHERE user_id=? AND status='draft'", (user_id,)).fetchall()
        for r in rows:
            if r["photo_path"] and os.path.exists(r["photo_path"]):
                try: os.remove(r["photo_path"])
                except: pass
        conn.execute("DELETE FROM items WHERE user_id=? AND status='draft'", (user_id,))
        conn.commit()

# -------------------- FSM --------------------
class AddItem(StatesGroup):
    waiting_photo = State()
    waiting_link = State()
    waiting_color = State()
    waiting_size = State()
    waiting_qty = State()
    waiting_comment = State()
    after_add_menu = State()

class EditItem(StatesGroup):
    choosing_item = State()
    choosing_field = State()
    editing_value = State()
    after_edit_menu = State()

class SaveExcel(StatesGroup):
    confirming = State()
    asking_code = State()

# -------------------- Клавиатуры --------------------
def main_menu_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="📂 Создать Excel", callback_data="menu:create")
    kb.button(text="📜 Посмотреть архив", callback_data="menu:archive")
    kb.adjust(2)
    return kb.as_markup()

def add_edit_menu_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="➕ Добавить товар", callback_data="item:add")
    kb.button(text="✏ Редактировать товар", callback_data="item:edit")
    kb.button(text="🗑 Удалить товар", callback_data="item:delete")
    kb.button(text="✅ Завершить Excel", callback_data="item:finish")
    kb.adjust(2,2)
    return kb.as_markup()

def back_to_items_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="🔙 Назад", callback_data="item:back_menu")
    return kb.as_markup()

def finish_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="💾 Сохранить Excel", callback_data="finish:save")
    kb.button(text="🔙 Назад", callback_data="finish:back")
    kb.adjust(2)
    return kb.as_markup()

def items_list_kb(items: List[sqlite3.Row], prefix: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    if not items:
        kb.button(text="Нет товаров", callback_data="noop")
    else:
        for it in items:
            title = f"{it['idx']}. {short_item_title(it)}"
            kb.button(text=title[:64], callback_data=f"{prefix}:{it['idx']}")
    kb.button(text="🔙 Назад", callback_data="item:back_menu")
    kb.adjust(1)
    return kb.as_markup()

def fields_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    for f, label in [
        ("photo", "Фото"), ("link", "Ссылка"), ("color", "Цвет"),
        ("size", "Размер"), ("qty", "Количество"), ("comment", "Комментарий")
    ]:
        kb.button(text=label, callback_data=f"field:{f}")
    kb.button(text="🔙 Назад", callback_data="item:edit")
    kb.adjust(2,2,2,1)
    return kb.as_markup()

def short_item_title(it: sqlite3.Row) -> str:
    parts = []
    if it["link"]: parts.append("🔗")
    if it["color"]: parts.append(it["color"])
    if it["size"]: parts.append(str(it["size"]))
    if it["qty"]: parts.append(f"x{it['qty']}")
    if not parts: parts.append("(пусто)")
    return " ".join(parts)

# -------------------- Генерация Excel --------------------
def build_excel_for_user(user_id: int, cargo_code: str) -> str:
    items = read_draft_items(user_id)
    archive_dir = ensure_archive_folder(user_id)
    safe_code = "".join(c for c in cargo_code if c.isalnum() or c in ("-","_"))
    if not safe_code:
        safe_code = f"cargo_{int(datetime.utcnow().timestamp())}"
    xlsx_path = os.path.join(archive_dir, f"{safe_code}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "MAGEZ"

    # Шапка (1-2 строки), красный фон, белый жирный текст
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")

    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws["A1"] = "MAGEZ"
    ws["A2"] = "Торгово-Логистическая компания"
    for cell in ("A1","A2"):
        c = ws[cell]
        c.fill = red
        c.font = white_bold
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # Заголовки с 4-й строки
    headers = ["Фото","Ссылка","Цвет","Размер","Количество","Комментарий"]
    start_row = 4
    ws.append([])  # row 3 пустая
    ws.append(headers)  # это будет 4-я строка

    gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 7):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = gray
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Данные
    row = start_row + 1
    col_widths = [len(h) for h in headers]
    for it in items:
        # Фото (колонка 1)
        photo_cell = ws.cell(row=row, column=1, value=None)
        photo_cell.alignment = center
        photo_cell.border = border
        if it["photo_path"] and os.path.exists(it["photo_path"]):
            try:
                # Пытаемся вставить изображение
                img = XLImage(it["photo_path"])
                # уменьшаем высоту строки под миниатюру
                ws.row_dimensions[row].height = 80
                # Вставляем в ячейку A{row}
                img.anchor = f"A{row}"
                ws.add_image(img)
                col_widths[0] = max(col_widths[0], 12)
            except Exception:
                # не похоже на картинку — пишем имя файла
                name = os.path.basename(it["photo_path"])
                photo_cell.value = name
                col_widths[0] = max(col_widths[0], len(name))
        else:
            photo_cell.value = ""
        # Остальные поля
        values = [it["link"] or "", it["color"] or "", it["size"] or "", it["qty"] or "", it["comment"] or ""]
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=row, column=i, value=v)
            cell.alignment = center
            cell.border = border
            col_widths[i-1] = max(col_widths[i-1], len(str(v)))
        row += 1

    # Авто-ширина (приблизительно)
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = max(10, min(50, w + 4))

    wb.save(xlsx_path)
    return xlsx_path

# -------------------- Бот и роутеры --------------------
router = Router()

def access_denied_text() -> str:
    return textwrap.dedent("""\
    🔒 Доступ к боту по пригласительным кодам.
    Отправьте /start <код_приглашения> чтобы войти.
    """)

async def show_main_menu(message_or_cb):
    if isinstance(message_or_cb, Message):
        await message_or_cb.answer("Главное меню:", reply_markup=main_menu_kb())
    else:
        await message_or_cb.message.edit_text("Главное меню:", reply_markup=main_menu_kb())

# -------------------- Обработчики --------------------
@router.message(Command("start"))
async def cmd_start(message: Message, command: CommandObject, state: FSMContext):
    tg_id = message.from_user.id
    arg = (command.args or "").strip()

    # Если уже зарегистрирован — сразу меню
    if is_registered(tg_id):
        await show_main_menu(message)
        return

    # Если нет — требуется код
    if not arg:
        await message.answer(access_denied_text())
        return

    # Пробуем активировать код
    if use_invite(arg, tg_id):
        register_user(tg_id)
        await message.answer("✅ Добро пожаловать! Доступ предоставлен.")
        await show_main_menu(message)
    else:
        await message.answer("❌ Неверный или уже использованный код приглашения.\n" + access_denied_text())

@router.message(Command("invite"))
async def cmd_invite(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return await message.answer("Команда доступна только админам.")
    code = create_invite(message.from_user.id)
    await message.answer(f"Создан одноразовый код:\n`{code}`\nОтправьте новому пользователю: /start {code}", parse_mode="Markdown")

@router.message(Command("invites"))
async def cmd_invites(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return await message.answer("Команда доступна только админам.")
    rows = list_open_invites()
    if not rows:
        return await message.answer("Открытых пригласительных нет.")
    txt = "Открытые коды:\n" + "\n".join(f"• `{r['code']}` (создан {r['created_at']})" for r in rows)
    await message.answer(txt, parse_mode="Markdown")

# Главный экран (inline-кнопки)
@router.callback_query(F.data == "menu:create")
async def menu_create(cb: CallbackQuery, state: FSMContext):
    if not is_registered(cb.from_user.id):
        return await cb.answer("Нет доступа", show_alert=True)
    await state.clear()
    await cb.message.edit_text("Создание Excel. Что дальше?", reply_markup=add_edit_menu_kb())

@router.callback_query(F.data == "menu:archive")
async def menu_archive(cb: CallbackQuery):
    if not is_registered(cb.from_user.id):
        return await cb.answer("Нет доступа", show_alert=True)
    user_dir = ensure_archive_folder(cb.from_user.id)
    files = sorted([f for f in os.listdir(user_dir) if f.lower().endswith(".xlsx")])
    if not files:
        return await cb.message.edit_text("📭 В архиве файлов нет.", reply_markup=main_menu_kb())
    kb = InlineKeyboardBuilder()
    for f in files:
        kb.button(text=f, callback_data=f"openfile:{f}")
    kb.button(text="🔙 Назад", callback_data="back:main")
    kb.adjust(1)
    await cb.message.edit_text("📜 Ваш архив:", reply_markup=kb.as_markup())

@router.callback_query(F.data == "back:main")
async def back_main(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await show_main_menu(cb)

@router.callback_query(F.data.startswith("openfile:"))
async def openfile(cb: CallbackQuery):
    if not is_registered(cb.from_user.id):
        return await cb.answer("Нет доступа", show_alert=True)
    fname = cb.data.split(":",1)[1]
    path = os.path.join(ensure_archive_folder(cb.from_user.id), fname)
    if not os.path.exists(path):
        return await cb.answer("Файл не найден", show_alert=True)
    await cb.message.answer_document(document=path, caption=f"📎 {fname}")
    await cb.answer()

# -------- Добавление товара --------
@router.callback_query(F.data == "item:add")
async def item_add_start(cb: CallbackQuery, state: FSMContext):
    await state.update_data(current_idx=next_item_index(cb.from_user.id))
    await state.set_state(AddItem.waiting_photo)
    await cb.message.edit_text("Отправьте фото товара (можно любое вложение).")

@router.message(AddItem.waiting_photo, F.content_type.in_({ContentType.PHOTO, ContentType.DOCUMENT, ContentType.ANIMATION, ContentType.VIDEO, ContentType.VIDEO_NOTE, ContentType.VOICE, ContentType.AUDIO}))
async def item_photo_received(message: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    idx = data["current_idx"]
    user_dir = ensure_user_folder(message.from_user.id)
    # Сохраняем файл
    filename = f"{idx}_{int(datetime.utcnow().timestamp())}"
    # пытаемся угадать расширение
    ext = ".bin"
    file_to_download = None
    if message.photo:
        file_to_download = message.photo[-1]
        ext = ".jpg"
    elif message.document:
        file_to_download = message.document
        name = message.document.file_name or ""
        if "." in name:
            ext = "." + name.rsplit(".",1)[1]
    else:
        file_to_download = message.content_type  # fallback
    path = os.path.join(user_dir, filename + ext)
    try:
        await message.bot.download(file_to_download, destination=path)
    except Exception:
        # не удалось — просто сохраним как бинарь
        with open(path,"wb") as f:
            f.write(message.model_dump_json().encode("utf-8", errors="ignore"))
    upsert_item_field(message.from_user.id, idx, "photo_path", path)
    await state.set_state(AddItem.waiting_link)
    await message.answer("Фото сохранено. Отправьте ссылку на товар (или '-' чтобы пропустить).")

@router.message(AddItem.waiting_link)
async def item_link(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = data["current_idx"]
    val = message.text.strip()
    upsert_item_field(message.from_user.id, idx, "link", "" if val == "-" else val)
    await state.set_state(AddItem.waiting_color)
    await message.answer("Цвет:")

@router.message(AddItem.waiting_color)
async def item_color(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = data["current_idx"]
    upsert_item_field(message.from_user.id, idx, "color", message.text.strip())
    await state.set_state(AddItem.waiting_size)
    await message.answer("Размер:")

@router.message(AddItem.waiting_size)
async def item_size(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = data["current_idx"]
    upsert_item_field(message.from_user.id, idx, "size", message.text.strip())
    await state.set_state(AddItem.waiting_qty)
    await message.answer("Количество:")

@router.message(AddItem.waiting_qty)
async def item_qty(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = data["current_idx"]
    upsert_item_field(message.from_user.id, idx, "qty", message.text.strip())
    await state.set_state(AddItem.waiting_comment)
    await message.answer("Комментарий (или '-' чтобы пропустить):")

@router.message(AddItem.waiting_comment)
async def item_comment(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = data["current_idx"]
    val = message.text.strip()
    upsert_item_field(message.from_user.id, idx, "comment", "" if val == "-" else val)
    await state.set_state(AddItem.after_add_menu)
    await message.answer("Товар добавлен. Что дальше?", reply_markup=add_edit_menu_kb())

# -------- Редактирование товара --------
@router.callback_query(F.data == "item:edit")
async def item_edit_start(cb: CallbackQuery, state: FSMContext):
    await state.set_state(EditItem.choosing_item)
    items = read_draft_items(cb.from_user.id)
    await cb.message.edit_text("Выберите товар для редактирования:", reply_markup=items_list_kb(items, "editpick"))

@router.callback_query(EditItem.choosing_item, F.data.startswith("editpick:"))
async def item_edit_pick(cb: CallbackQuery, state: FSMContext):
    idx = int(cb.data.split(":")[1])
    await state.update_data(edit_idx=idx)
    await state.set_state(EditItem.choosing_field)
    it = get_item_by_idx(cb.from_user.id, idx)
    text = f"Товар {idx}:\n{short_item_title(it)}\nВыберите поле для редактирования:"
    await cb.message.edit_text(text, reply_markup=fields_kb())

@router.callback_query(EditItem.choosing_field, F.data.startswith("field:"))
async def item_edit_field(cb: CallbackQuery, state: FSMContext):
    field = cb.data.split(":")[1]
    await state.update_data(edit_field=field)
    if field == "photo":
        await state.set_state(EditItem.editing_value)  # будем ждать вложение
        await cb.message.edit_text("Отправьте новое фото/файл:", reply_markup=back_to_items_kb())
    else:
        await state.set_state(EditItem.editing_value)
        await cb.message.edit_text(f"Введите новое значение '{field}':", reply_markup=back_to_items_kb())

@router.message(EditItem.editing_value, F.text, ~F.text.in_({"", None}))
async def item_edit_value_text(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = int(data["edit_idx"])
    field = data["edit_field"]
    if field == "photo":
        await message.answer("Это поле изменяется файлом. Отправьте фото/файл.")
        return
    upsert_item_field(message.from_user.id, idx, field if field != "photo" else "photo_path", message.text.strip())
    await state.set_state(EditItem.choosing_field)
    await message.answer("Изменено. Что ещё правим?", reply_markup=fields_kb())

@router.message(EditItem.editing_value, F.content_type.in_({ContentType.PHOTO, ContentType.DOCUMENT}))
async def item_edit_value_file(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = int(data["edit_idx"])
    field = data["edit_field"]
    if field != "photo":
        await message.answer("Ожидался текст. Введите новое значение.")
        return
    user_dir = ensure_user_folder(message.from_user.id)
    filename = f"{idx}_{int(datetime.utcnow().timestamp())}"
    ext = ".jpg" if message.photo else ".bin"
    if message.document and message.document.file_name and "." in message.document.file_name:
        ext = "." + message.document.file_name.rsplit(".",1)[1]
    path = os.path.join(user_dir, filename + ext)
    try:
        await message.bot.download(message.photo[-1] if message.photo else message.document, destination=path)
    except Exception:
        with open(path,"wb") as f:
            f.write(message.model_dump_json().encode("utf-8", errors="ignore"))
    upsert_item_field(message.from_user.id, idx, "photo_path", path)
    await state.set_state(EditItem.choosing_field)
    await message.answer("Фото обновлено. Что ещё правим?", reply_markup=fields_kb())

# -------- Удаление товара --------
@router.callback_query(F.data == "item:delete")
async def item_delete_start(cb: CallbackQuery, state: FSMContext):
    await state.set_state(EditItem.choosing_item)
    items = read_draft_items(cb.from_user.id)
    await cb.message.edit_text("Выберите товар для удаления:", reply_markup=items_list_kb(items, "delpick"))

@router.callback_query(EditItem.choosing_item, F.data.startswith("delpick:"))
async def item_delete_confirm(cb: CallbackQuery, state: FSMContext):
    idx = int(cb.data.split(":")[1])
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Да, удалить", callback_data=f"delconfirm:{idx}")
    kb.button(text="❌ Отмена", callback_data="item:back_menu")
    kb.adjust(1)
    it = get_item_by_idx(cb.from_user.id, idx)
    await cb.message.edit_text(f"Удалить товар {idx}: {short_item_title(it)} ?", reply_markup=kb.as_markup())

@router.callback_query(F.data.startswith("delconfirm:"))
async def item_delete_do(cb: CallbackQuery, state: FSMContext):
    idx = int(cb.data.split(":")[1])
    delete_item(cb.from_user.id, idx)
    await cb.message.edit_text("Удалено. Что дальше?", reply_markup=add_edit_menu_kb())

# -------- Завершение Excel --------
@router.callback_query(F.data == "item:finish")
async def item_finish(cb: CallbackQuery, state: FSMContext):
    items = read_draft_items(cb.from_user.id)
    if not items:
        return await cb.answer("Нет товаров для сохранения.", show_alert=True)
    # Показать список
    text_lines = ["Список товаров:"]
    for it in items:
        text_lines.append(f"{it['idx']}. {short_item_title(it)}")
    await state.set_state(SaveExcel.confirming)
    await cb.message.edit_text("\n".join(text_lines), reply_markup=finish_kb())

@router.callback_query(SaveExcel.confirming, F.data == "finish:back")
async def finish_back(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.edit_text("Возвращаюсь к редактированию:", reply_markup=add_edit_menu_kb())

@router.callback_query(SaveExcel.confirming, F.data == "finish:save")
async def finish_save(cb: CallbackQuery, state: FSMContext):
    await state.set_state(SaveExcel.asking_code)
    await cb.message.edit_text("Введите код груза (имя файла Excel):", reply_markup=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="🔙 Назад", callback_data="finish:back")]]))

@router.message(SaveExcel.asking_code)
async def finish_code_entered(message: Message, state: FSMContext):
    cargo_code = message.text.strip()
    try:
        xlsx_path = build_excel_for_user(message.from_user.id, cargo_code)
    except Exception as e:
        await message.answer(f"Ошибка при сохранении Excel: {e}")
        return
    # Отправляем файл
    await message.answer_document(document=xlsx_path, caption=f"💾 Файл сохранён: {os.path.basename(xlsx_path)}")
    # Чистим черновик
    reset_draft(message.from_user.id)
    await state.clear()
    await message.answer("Готово. Возвращаю в главное меню.", reply_markup=main_menu_kb())

# -------- Навигация назад к меню Add/Edit --------
@router.callback_query(F.data == "item:back_menu")
async def back_items_menu(cb: CallbackQuery, state: FSMContext):
    await state.clear()
    await cb.message.edit_text("Меню создания Excel:", reply_markup=add_edit_menu_kb())

# -------- Заглушка для noop --------
@router.callback_query(F.data == "noop")
async def noop(cb: CallbackQuery):
    await cb.answer()

# -------- Защита: все прочие сообщения без регистрации --------
@router.message()
async def fallback(message: Message):
    if not is_registered(message.from_user.id):
        await message.answer(access_denied_text())
    else:
        await message.answer("Используйте кнопки выше.", reply_markup=main_menu_kb())

# -------------------- main --------------------
async def main():
    if not BOT_TOKEN:
        raise SystemExit("Не задан BOT_TOKEN в переменных окружения.")
    init_db()
    bot = Bot(BOT_TOKEN, parse_mode="HTML")
    dp = Dispatcher()
    dp.include_router(router)
    print("Bot is up. Press Ctrl+C to stop.")
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        print("Bot stopped.")
