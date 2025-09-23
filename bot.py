# bot.py
# -*- coding: utf-8 -*-
"""
Telegram бот (aiogram 3.x) с одноразовыми приглашениями, главным меню (inline),
FSM для добавления/редактирования/удаления товаров и экспортом в Excel с нужным оформлением.
Под Replit и GitHub. Максимально в одном файле (этот), вспомогательные данные — в JSON/папках.

Переменные окружения:
- TELEGRAM_BOT_TOKEN=<ваш токен из BotFather>
- ADMIN_IDS=123456789,987654321   (список Telegram ID админов через запятую)

Запуск локально:
    pip install -r requirements.txt
    python bot.py

Автор: ваш наставник по Python :)
"""
import asyncio
import os
import json
import io
import math
import textwrap
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from aiogram import Bot, Dispatcher, F, types
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import InlineKeyboardBuilder

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, UnidentifiedImageError

# ---------- Константы/пути ----------
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
TMP_DIR = DATA_DIR / "tmp"
INVITES_FILE = DATA_DIR / "invites.json"
USERS_FILE = DATA_DIR / "authorized_users.json"
ARCHIVE_DIR = BASE_DIR / "archive"

# Создаём необходимые папки
for p in [DATA_DIR, TMP_DIR, ARCHIVE_DIR]:
    p.mkdir(parents=True, exist_ok=True)

# ---------- Утилиты JSON ----------
def load_json(path: Path, default):
    if not path.exists():
        return default
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default

def save_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ---------- Приглашения и доступ ----------
def get_admin_ids() -> List[int]:
    raw = os.getenv("ADMIN_IDS", "").strip()
    ids = []
    for part in raw.split(","):
        part = part.strip()
        if part.isdigit():
            ids.append(int(part))
    return ids

def is_admin(user_id: int) -> bool:
    return user_id in get_admin_ids()

def list_invites() -> List[Dict[str, Any]]:
    return load_json(INVITES_FILE, [])

def save_invites(invites: List[Dict[str, Any]]):
    save_json(INVITES_FILE, invites)

def create_invite_codes(count: int) -> List[str]:
    import secrets
    invites = list_invites()
    existing = {item["code"] for item in invites}
    new_codes = []
    while len(new_codes) < count:
        code = secrets.token_urlsafe(8).replace("_", "").replace("-", "")
        if code not in existing:
            new_codes.append(code)
            invites.append({"code": code, "used": False, "used_by": None, "used_at": None})
            existing.add(code)
    save_invites(invites)
    return new_codes

def mark_invite_used(code: str, user_id: int):
    invites = list_invites()
    changed = False
    for item in invites:
        if item["code"] == code and not item.get("used"):
            item["used"] = True
            item["used_by"] = user_id
            item["used_at"] = datetime.utcnow().isoformat()
            changed = True
            break
    if changed:
        save_invites(invites)

def is_valid_invite(code: str) -> bool:
    for item in list_invites():
        if item["code"] == code and not item.get("used"):
            return True
    return False

def list_authorized_users() -> List[int]:
    return load_json(USERS_FILE, [])

def save_authorized_users(uids: List[int]):
    save_json(USERS_FILE, uids)

def is_authorized(user_id: int) -> bool:
    return user_id in list_authorized_users()

def authorize_user(user_id: int):
    users = list_authorized_users()
    if user_id not in users:
        users.append(user_id)
        save_authorized_users(users)

# ---------- Модель данных товара ----------
class Item:
    def __init__(self):
        self.photo_path: Optional[str] = None  # путь к сохранённому файлу
        self.link: Optional[str] = None
        self.color: Optional[str] = None
        self.size: Optional[str] = None
        self.qty: Optional[int] = None
        self.comment: Optional[str] = None

    def to_dict(self):
        return {
            "photo_path": self.photo_path,
            "link": self.link,
            "color": self.color,
            "size": self.size,
            "qty": self.qty,
            "comment": self.comment
        }

    @classmethod
    def from_dict(cls, d: Dict[str, Any]):
        it = cls()
        it.photo_path = d.get("photo_path")
        it.link = d.get("link")
        it.color = d.get("color")
        it.size = d.get("size")
        it.qty = d.get("qty")
        it.comment = d.get("comment")
        return it

    def short(self) -> str:
        parts = []
        if self.link: parts.append(f"🔗 {self.link}")
        if self.color: parts.append(f"🎨 {self.color}")
        if self.size: parts.append(f"📏 {self.size}")
        if self.qty is not None: parts.append(f"📦 {self.qty}")
        if self.comment: parts.append(f"💬 {self.comment}")
        if self.photo_path: parts.append(f"🖼 {Path(self.photo_path).name}")
        return " | ".join(parts) if parts else "(пусто)"

# ---------- FSM ----------
class AddItemStates(StatesGroup):
    waiting_photo = State()
    waiting_link = State()
    waiting_color = State()
    waiting_size = State()
    waiting_qty = State()
    waiting_comment = State()

class EditFieldStates(StatesGroup):
    choosing_item = State()
    choosing_field = State()
    editing_value = State()

class SaveExcelStates(StatesGroup):
    confirm_list = State()
    waiting_cargo_code = State()

# ---------- Хранилище сеансов (в памяти процесса) ----------
# Для простоты — в памяти. В проде лучше БД.
USER_ITEMS: Dict[int, List[Item]] = {}     # user_id -> список товаров
EDIT_CONTEXT: Dict[int, Dict[str, Any]] = {}  # user_id -> {"index": int, "field": str}

# ---------- Клавиатуры ----------
def main_menu_kb() -> types.InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="📂 Создать Excel", callback_data="create_excel")
    kb.button(text="📜 Посмотреть архив", callback_data="view_archive")
    kb.adjust(1)
    return kb.as_markup()

def post_add_menu_kb() -> types.InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="➕ Добавить товар", callback_data="add_item")
    kb.button(text="✏ Редактировать товар", callback_data="edit_item")
    kb.button(text="🗑 Удалить товар", callback_data="delete_item")
    kb.button(text="✅ Завершить Excel", callback_data="finalize_excel")
    kb.adjust(1)
    return kb.as_markup()

def items_list_kb(user_id: int, action_prefix: str) -> types.InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    items = USER_ITEMS.get(user_id, [])
    if not items:
        kb.button(text="(список пуст)", callback_data="noop")
    else:
        for idx, it in enumerate(items):
            label = f"{idx+1}. {it.short()[:64]}"
            kb.button(text=label, callback_data=f"{action_prefix}:{idx}")
    kb.button(text="🔙 Назад", callback_data="back_to_post_add")
    kb.adjust(1)
    return kb.as_markup()

def edit_fields_kb() -> types.InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    for f, label in [
        ("photo", "Фото"),
        ("link", "Ссылка"),
        ("color", "Цвет"),
        ("size", "Размер"),
        ("qty", "Количество"),
        ("comment", "Комментарий"),
    ]:
        kb.button(text=label, callback_data=f"edit_field:{f}")
    kb.button(text="🔙 Назад", callback_data="back_choose_item")
    kb.adjust(2)
    return kb.as_markup()

def save_or_back_kb() -> types.InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="💾 Сохранить Excel", callback_data="save_excel")
    kb.button(text="🔙 Назад", callback_data="back_to_post_add")
    kb.adjust(1)
    return kb.as_markup()

# ---------- Помощники ----------
async def ensure_authorized(event: types.TelegramObject, bot: Bot) -> bool:
    """Проверка доступа. Если нет — отправим подсказку. Возвращает True/False."""
    user = event.from_user
    if user is None:
        return False
    if is_authorized(user.id) or is_admin(user.id):
        return True
    # Неавторизован
    text = ("Доступ только по приглашению.\n"
            "Попросите у администратора одноразовый код и используйте: /start <код>")
    if isinstance(event, types.Message):
        await event.answer(text)
    elif isinstance(event, types.CallbackQuery):
        await event.message.answer(text)
    return False

def get_user_tmp_dir(user_id: int) -> Path:
    p = TMP_DIR / str(user_id)
    p.mkdir(parents=True, exist_ok=True)
    return p

def get_user_archive_dir(user_id: int) -> Path:
    p = ARCHIVE_DIR / str(user_id)
    p.mkdir(parents=True, exist_ok=True)
    return p

def clear_user_tmp(user_id: int):
    p = get_user_tmp_dir(user_id)
    for child in p.glob("*"):
        try:
            if child.is_file():
                child.unlink()
        except Exception:
            pass

def human_file_list(dir_path: Path) -> List[Path]:
    return sorted([p for p in dir_path.glob("*.xlsx")], key=lambda x: x.stat().st_mtime, reverse=True)

def center_alignment():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def border_thin():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def fit_image_to_cell(img_path: Path, target_w_px: int, target_h_px: int) -> PILImage:
    img = PILImage.open(img_path)
    img.thumbnail((target_w_px, target_h_px), PILImage.LANCZOS)
    return img

def pixels_to_excel_width(px: int) -> float:
    # Приблизительное преобразование px -> column width (экспериментально)
    # Ширина 1 = ~7 пикселей текста Calibri 11.
    return max(10.0, px / 7.0)

def best_col_width(values: List[str], min_width: float = 12.0, max_width: float = 60.0) -> float:
    # Примерно оцениваем ширину по длине текста
    max_len = max((len(v) for v in values if v), default=0)
    width = max(min_width, min(max_width, max_len * 1.2))
    return width

# ---------- Excel генерация ----------
def build_excel(user_id: int, cargo_code: str, items: List[Item]) -> Path:
    user_dir = get_user_archive_dir(user_id)
    safe_code = "".join([c for c in cargo_code if c.isalnum() or c in ("-", "_")]) or f"cargo_{datetime.now():%Y%m%d_%H%M%S}"
    filename = f"{safe_code}.xlsx"
    out_path = user_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказ"

    # Настройка шапки (строки 1-2, красный фон, белый жирный текст, объединение)
    header_fill = PatternFill("solid", fgColor="FF0000")  # красный
    header_font = Font(b=True, color="FFFFFF", size=14)

    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws["A1"] = "MAGEZ"
    ws["A2"] = "Торгово-Логистическая компания"
    for row in (1, 2):
        cell = ws[f"A{row}"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment()
        ws.row_dimensions[row].height = 24

    # Заголовки таблицы с 4-й строки
    start_row = 4
    headers = ["Фото", "Ссылка", "Цвет", "Размер", "Количество", "Комментарий"]
    header_fill_gray = PatternFill("solid", fgColor="808080")
    header_font_white = Font(b=True, color="FFFFFF")

    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=name)
        cell.fill = header_fill_gray
        cell.font = header_font_white
        cell.alignment = center_alignment()
        cell.border = border_thin()
    ws.row_dimensions[start_row].height = 24

    # Данные
    row = start_row + 1
    # Параметры изображений
    # Примем высоту строки 100 px для фото; Excel высота ~ 0.75 * px
    photo_row_height_px = 120
    photo_row_height_points = photo_row_height_px * 0.75
    ws.column_dimensions["A"].width = pixels_to_excel_width(140)

    # Соберём значения для оценки ширины столбцов
    col_values = {i: [] for i in range(1, 7)}
    for it in items:
        col_values[2].append(it.link or "")
        col_values[3].append(it.color or "")
        col_values[4].append(it.size or "")
        col_values[5].append(str(it.qty) if it.qty is not None else "")
        col_values[6].append(it.comment or "")

    for it in items:
        ws.row_dimensions[row].height = photo_row_height_points
        # Фото / файл
        if it.photo_path and Path(it.photo_path).exists():
            img_path = Path(it.photo_path)
            placed = False
            try:
                # Подгоним по размеру ячейки (примерно)
                target_w_px = 140
                target_h_px = photo_row_height_px
                pil_img = fit_image_to_cell(img_path, target_w_px, target_h_px)
                buf = io.BytesIO()
                pil_img.save(buf, format="PNG")
                buf.seek(0)
                xl_img = XLImage(buf)
                cell_addr = f"A{row}"
                ws.add_image(xl_img, cell_addr)
                # Выравнивание и границы ячейки тоже поставим
                cell = ws.cell(row=row, column=1, value=None)
                cell.alignment = center_alignment()
                cell.border = border_thin()
                placed = True
            except UnidentifiedImageError:
                placed = False
            if not placed:
                # не изображение — пишем имя файла
                cell = ws.cell(row=row, column=1, value=img_path.name)
                cell.alignment = center_alignment()
                cell.border = border_thin()
        else:
            cell = ws.cell(row=row, column=1, value="—")
            cell.alignment = center_alignment()
            cell.border = border_thin()

        # Остальные поля
        vals = [it.link, it.color, it.size, it.qty, it.comment]
        for i, v in enumerate(vals, start=2):
            cell = ws.cell(row=row, column=i, value=v if v is not None else "")
            cell.alignment = center_alignment()
            cell.border = border_thin()
        row += 1

    # Авто-ширина (примерная)
    ws.column_dimensions["B"].width = best_col_width(col_values[2], 20, 80)
    ws.column_dimensions["C"].width = best_col_width(col_values[3], 14, 40)
    ws.column_dimensions["D"].width = best_col_width(col_values[4], 14, 30)
    ws.column_dimensions["E"].width = best_col_width([str(x) for x in col_values[5]], 12, 16)
    ws.column_dimensions["F"].width = best_col_width(col_values[6], 20, 80)

    wb.save(out_path)
    return out_path

# ---------- Инициализация бота ----------
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not TOKEN:
    raise SystemExit("Не найден TELEGRAM_BOT_TOKEN в переменных окружения.")

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

# ---------- Хэндлеры ----------

# /start с опциональным кодом приглашения
@dp.message(CommandStart(deep_link=True))
async def start_deeplink(message: types.Message):
    code = message.text.split(maxsplit=1)
    code = code[1].strip() if len(code) > 1 else ""
    user_id = message.from_user.id

    if is_authorized(user_id) or is_admin(user_id):
        await message.answer("С возвращением! Главное меню:", reply_markup=main_menu_kb())
        return

    if code and is_valid_invite(code):
        authorize_user(user_id)
        mark_invite_used(code, user_id)
        await message.answer("Вы успешно авторизованы по приглашению.\nГлавное меню:", reply_markup=main_menu_kb())
    else:
        await message.answer(
            "Доступ только по приглашению.\n"
            "Получите одноразовый код у администратора и используйте: <code>/start КОД</code>"
        )

# /start без кода
@dp.message(CommandStart())
async def start(message: types.Message):
    if not await ensure_authorized(message, bot):
        return
    await message.answer("Главное меню:", reply_markup=main_menu_kb())

# Команда для админа: создание приглашений
@dp.message(Command("new_invite"))
async def cmd_new_invite(message: types.Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("Команда доступна только администратору.")
        return
    # /new_invite 3  -> создать 3 кода
    parts = message.text.split()
    count = 1
    if len(parts) > 1 and parts[1].isdigit():
        count = max(1, min(50, int(parts[1])))
    codes = create_invite_codes(count)
    text = "Созданы приглашения:\n" + "\n".join([f"<code>{c}</code>" for c in codes])
    await message.answer(text)

# Главное меню: кнопки
@dp.callback_query(F.data == "create_excel")
async def cb_create_excel(callback: types.CallbackQuery, state: FSMContext):
    if not await ensure_authorized(callback, bot):
        return
    user_id = callback.from_user.id
    USER_ITEMS.setdefault(user_id, [])
    clear_user_tmp(user_id)
    await callback.message.edit_text(
        "Добавьте товары. После каждого добавления доступны действия ниже.",
        reply_markup=post_add_menu_kb()
    )

@dp.callback_query(F.data == "view_archive")
async def cb_view_archive(callback: types.CallbackQuery):
    if not await ensure_authorized(callback, bot):
        return
    user_id = callback.from_user.id
    files = human_file_list(get_user_archive_dir(user_id))
    if not files:
        await callback.message.edit_text("В вашем архиве пока нет файлов.", reply_markup=main_menu_kb())
        return

    # Показать список файлов с кнопками для отправки
    kb = InlineKeyboardBuilder()
    for f in files[:50]:
        kb.button(text=f.name, callback_data=f"send_file:{f.name}")
    kb.button(text="🔙 В меню", callback_data="back_to_menu")
    kb.adjust(1)
    await callback.message.edit_text("Ваш архив:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("send_file:"))
async def cb_send_file(callback: types.CallbackQuery):
    if not await ensure_authorized(callback, bot):
        return
    user_id = callback.from_user.id
    name = callback.data.split(":", 1)[1]
    path = get_user_archive_dir(user_id) / name
    if not path.exists():
        await callback.answer("Файл не найден.", show_alert=True)
        return
    await callback.message.answer_document(types.FSInputFile(path))
    await callback.answer()

@dp.callback_query(F.data == "back_to_menu")
async def cb_back_to_menu(callback: types.CallbackQuery):
    await callback.message.edit_text("Главное меню:", reply_markup=main_menu_kb())

# --- Пост-меню (после создания Excel) ---
@dp.callback_query(F.data == "add_item")
async def cb_add_item(callback: types.CallbackQuery, state: FSMContext):
    if not await ensure_authorized(callback, bot):
        return
    await state.set_state(AddItemStates.waiting_photo)
    await callback.message.edit_text("Шаг 1/6 — пришлите Фото (можно фото или документ).")

@dp.message(AddItemStates.waiting_photo, F.photo | F.document)
async def add_photo(message: types.Message, state: FSMContext):
    if not await ensure_authorized(message, bot):
        return
    user_id = message.from_user.id
    tmp_dir = get_user_tmp_dir(user_id)

    file_id = None
    filename = None
    is_image = False

    if message.photo:
        # Берём самое большое фото
        file_id = message.photo[-1].file_id
        filename = f"photo_{message.photo[-1].file_unique_id}.jpg"
        is_image = True
    elif message.document:
        file_id = message.document.file_id
        filename = message.document.file_name or f"file_{message.document.file_unique_id}"
        # Пытаемся понять, похоже ли на изображение
        if (message.document.mime_type or "").startswith("image/"):
            is_image = True

    if not file_id:
        await message.answer("Не удалось получить файл, пришлите ещё раз.")
        return

    save_path = tmp_dir / filename
    try:
        await bot.download(file_id, destination=save_path)
    except Exception as e:
        await message.answer(f"Ошибка сохранения файла: {e}")
        return

    # Создаём заготовку Item и держим её во временном состоянии
    await state.update_data(photo_path=str(save_path))
    await state.set_state(AddItemStates.waiting_link)
    await message.answer("Шаг 2/6 — отправьте ссылку.")

@dp.message(AddItemStates.waiting_link)
async def add_link(message: types.Message, state: FSMContext):
    if not await ensure_authorized(message, bot):
        return
    await state.update_data(link=message.text.strip())
    await state.set_state(AddItemStates.waiting_color)
    await message.answer("Шаг 3/6 — укажите цвет.")

@dp.message(AddItemStates.waiting_color)
async def add_color(message: types.Message, state: FSMContext):
    if not await ensure_authorized(message, bot):
        return
    await state.update_data(color=message.text.strip())
    await state.set_state(AddItemStates.waiting_size)
    await message.answer("Шаг 4/6 — укажите размер.")

@dp.message(AddItemStates.waiting_size)
async def add_size(message: types.Message, state: FSMContext):
    if not await ensure_authorized(message, bot):
        return
    await state.update_data(size=message.text.strip())
    await state.set_state(AddItemStates.waiting_qty)
    await message.answer("Шаг 5/6 — укажите количество (число).")

@dp.message(AddItemStates.waiting_qty)
async def add_qty(message: types.Message, state: FSMContext):
    if not await ensure_authorized(message, bot):
        return
    txt = (message.text or "").strip()
    try:
        qty = int(txt)
    except ValueError:
        await message.answer("Введите целое число для количества.")
        return
    await state.update_data(qty=qty)
    await state.set_state(AddItemStates.waiting_comment)
    await message.answer("Шаг 6/6 — добавьте комментарий (или - для пропуска).")

@dp.message(AddItemStates.waiting_comment)
async def add_comment(message: types.Message, state: FSMContext):
    if not await ensure_authorized(message, bot):
        return
    comment = message.text.strip()
    if comment == "-":
        comment = ""
    data = await state.get_data()
    it = Item()
    it.photo_path = data.get("photo_path")
    it.link = data.get("link")
    it.color = data.get("color")
    it.size = data.get("size")
    it.qty = data.get("qty")
    it.comment = comment

    user_id = message.from_user.id
    USER_ITEMS.setdefault(user_id, []).append(it)
    await state.clear()

    await message.answer(
        "Товар добавлен ✅",
    )
    await message.answer(
        "Выберите действие:",
        reply_markup=post_add_menu_kb()
    )

# --- Редактирование ---
@dp.callback_query(F.data == "edit_item")
async def cb_edit_item(callback: types.CallbackQuery, state: FSMContext):
    if not await ensure_authorized(callback, bot):
        return
    user_id = callback.from_user.id
    if not USER_ITEMS.get(user_id):
        await callback.answer("Список пуст.", show_alert=True)
        return
    await state.set_state(EditFieldStates.choosing_item)
    await callback.message.edit_text("Выберите товар для редактирования:", reply_markup=items_list_kb(user_id, "choose_to_edit"))

@dp.callback_query(EditFieldStates.choosing_item, F.data.startswith("choose_to_edit:"))
async def cb_choose_item_for_edit(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    idx_str = callback.data.split(":")[1]
    try:
        idx = int(idx_str)
    except:
        await callback.answer("Ошибка индекса.", show_alert=True)
        return
    EDIT_CONTEXT[user_id] = {"index": idx}
    await state.set_state(EditFieldStates.choosing_field)
    await callback.message.edit_text("Выберите поле для редактирования:", reply_markup=edit_fields_kb())

@dp.callback_query(EditFieldStates.choosing_field, F.data.startswith("edit_field:"))
async def cb_choose_field(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    field = callback.data.split(":")[1]
    ctx = EDIT_CONTEXT.get(user_id, {})
    if "index" not in ctx:
        await callback.answer("Не выбран товар.", show_alert=True)
        return
    ctx["field"] = field
    EDIT_CONTEXT[user_id] = ctx

    if field == "photo":
        await state.set_state(EditFieldStates.editing_value)
        await callback.message.edit_text("Пришлите новое фото/файл для товара.")
    elif field == "qty":
        await state.set_state(EditFieldStates.editing_value)
        await callback.message.edit_text("Введите новое количество (число).")
    else:
        await state.set_state(EditFieldStates.editing_value)
        await callback.message.edit_text(f"Введите новое значение для поля «{field}». (или - чтобы очистить)")

@dp.message(EditFieldStates.editing_value, F.photo | F.document)
async def edit_field_file(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    ctx = EDIT_CONTEXT.get(user_id, {})
    field = ctx.get("field")
    if field != "photo":
        await message.answer("Ожидалось текстовое значение.")
        return

    tmp_dir = get_user_tmp_dir(user_id)
    if message.photo:
        file_id = message.photo[-1].file_id
        filename = f"photo_{message.photo[-1].file_unique_id}.jpg"
    else:
        file_id = message.document.file_id
        filename = message.document.file_name or f"file_{message.document.file_unique_id}"

    save_path = tmp_dir / filename
    try:
        await bot.download(file_id, destination=save_path)
    except Exception as e:
        await message.answer(f"Ошибка сохранения файла: {e}")
        return

    items = USER_ITEMS.get(user_id, [])
    idx = ctx.get("index", -1)
    if 0 <= idx < len(items):
        items[idx].photo_path = str(save_path)
        await message.answer("Фото обновлено.")
    else:
        await message.answer("Элемент не найден.")

    await state.clear()
    await message.answer("Выберите действие:", reply_markup=post_add_menu_kb())

@dp.message(EditFieldStates.editing_value)
async def edit_field_text(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    ctx = EDIT_CONTEXT.get(user_id, {})
    field = ctx.get("field")
    items = USER_ITEMS.get(user_id, [])
    idx = ctx.get("index", -1)

    if not (0 <= idx < len(items)):
        await message.answer("Элемент не найден.")
        await state.clear()
        await message.answer("Выберите действие:", reply_markup=post_add_menu_kb())
        return

    value = message.text.strip()
    if value == "-":
        value = ""

    if field == "qty":
        try:
            value = int(value)
        except ValueError:
            await message.answer("Введите целое число для количества.")
            return

    if field in ("link", "color", "size", "qty", "comment"):
        setattr(items[idx], field if field != "qty" else "qty", value)
        await message.answer("Значение обновлено.")
    else:
        await message.answer("Неизвестное поле.")

    await state.clear()
    await message.answer("Выберите действие:", reply_markup=post_add_menu_kb())

# --- Удаление ---
@dp.callback_query(F.data == "delete_item")
async def cb_delete_item(callback: types.CallbackQuery):
    if not await ensure_authorized(callback, bot):
        return
    user_id = callback.from_user.id
    if not USER_ITEMS.get(user_id):
        await callback.answer("Список пуст.", show_alert=True)
        return
    await callback.message.edit_text("Выберите товар для удаления:", reply_markup=items_list_kb(user_id, "choose_to_delete"))

@dp.callback_query(F.data.startswith("choose_to_delete:"))
async def cb_choose_to_delete(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    idx_str = callback.data.split(":")[1]
    try:
        idx = int(idx_str)
    except:
        await callback.answer("Ошибка индекса.", show_alert=True)
        return
    # Подтверждение
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Да, удалить", callback_data=f"confirm_delete:{idx}")
    kb.button(text="❌ Отмена", callback_data="back_to_post_add")
    kb.adjust(1)
    await callback.message.edit_text(f"Удалить товар #{idx+1}?", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("confirm_delete:"))
async def cb_confirm_delete(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    idx = int(callback.data.split(":")[1])
    items = USER_ITEMS.get(user_id, [])
    if 0 <= idx < len(items):
        removed = items.pop(idx)
        await callback.message.edit_text("Товар удалён.", reply_markup=post_add_menu_kb())
    else:
        await callback.answer("Элемент не найден.", show_alert=True)

# --- Завершение Excel ---
@dp.callback_query(F.data == "finalize_excel")
async def cb_finalize_excel(callback: types.CallbackQuery, state: FSMContext):
    if not await ensure_authorized(callback, bot):
        return
    user_id = callback.from_user.id
    items = USER_ITEMS.get(user_id, [])
    if not items:
        await callback.answer("Список пуст.", show_alert=True)
        return
    # Показ списка
    lines = []
    for i, it in enumerate(items, 1):
        lines.append(f"<b>{i}.</b> {it.short()}")
    text = "Список товаров:\n" + "\n".join(lines)
    await state.set_state(SaveExcelStates.confirm_list)
    await callback.message.edit_text(text, reply_markup=save_or_back_kb())

@dp.callback_query(SaveExcelStates.confirm_list, F.data == "save_excel")
async def cb_save_excel_ask_code(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(SaveExcelStates.waiting_cargo_code)
    await callback.message.edit_text("Введите код груза (он будет именем файла).")

@dp.message(SaveExcelStates.waiting_cargo_code)
async def save_excel_do(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    code = (message.text or "").strip()
    if not code:
        await message.answer("Код не должен быть пустым. Введите код груза.")
        return
    items = USER_ITEMS.get(user_id, [])
    path = build_excel(user_id, code, items)

    # Отправка файла
    await message.answer_document(types.FSInputFile(path), caption=f"Сохранено в архив: {path.name}")
    # Очистка временных файлов и списка
    clear_user_tmp(user_id)
    USER_ITEMS[user_id] = []

    await state.clear()
    await message.answer("Готово! Возврат в главное меню.", reply_markup=main_menu_kb())

# --- Навигация назад ---
@dp.callback_query(F.data == "back_to_post_add")
async def cb_back_post_add(callback: types.CallbackQuery):
    await callback.message.edit_text("Выберите действие:", reply_markup=post_add_menu_kb())

@dp.callback_query(F.data == "back_choose_item")
async def cb_back_choose_item(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    await state.set_state(EditFieldStates.choosing_item)
    await callback.message.edit_text("Выберите товар для редактирования:", reply_markup=items_list_kb(user_id, "choose_to_edit"))

@dp.callback_query(F.data == "noop")
async def cb_noop(callback: types.CallbackQuery):
    await callback.answer()

# --- Фолбэк на любые другие сообщения в "меню" ---
@dp.message(Command("menu"))
async def cmd_menu(message: types.Message):
    if not await ensure_authorized(message, bot):
        return
    await message.answer("Главное меню:", reply_markup=main_menu_kb())

# ---------- Точка входа ----------
async def main():
    print("Bot started...")
    await dp.start_polling(bot)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        print("Bot stopped.")
