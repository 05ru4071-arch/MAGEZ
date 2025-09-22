import os
import asyncio
from io import BytesIO
from aiogram import Bot, Dispatcher, types
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 🔑 Токен
TOKEN = "7741928533:AAFDsO77wqRsWLTR7cu39UQDvqMc5MsyEw4"

bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# 📦 Хранилище товаров: user_id -> список словарей
user_data = {}

# FSM состояния
class ProductForm(StatesGroup):
    waiting_for_photo = State()
    waiting_for_link = State()
    waiting_for_color = State()
    waiting_for_size = State()
    waiting_for_qty = State()
    waiting_for_comment = State()

class SaveExcel(StatesGroup):
    waiting_for_code = State()

class EditProduct(StatesGroup):
    choosing_product = State()
    choosing_field = State()
    updating_value = State()

class DeleteProduct(StatesGroup):
    choosing_product = State()

# Главное меню
def main_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📂 Создать Excel", callback_data="create_excel")],
        [InlineKeyboardButton(text="📜 Посмотреть архив", callback_data="view_archive")]
    ])

# Меню работы с товарами
def product_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить товар", callback_data="add_item")],
        [InlineKeyboardButton(text="✏ Редактировать товар", callback_data="edit_item")],
        [InlineKeyboardButton(text="🗑 Удалить товар", callback_data="delete_item")],
        [InlineKeyboardButton(text="✅ Завершить Excel", callback_data="finish_excel")]
    ])

# Меню завершения
def finish_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💾 Сохранить Excel", callback_data="save_excel")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_products")]
    ])

# /start
@dp.message(Command("start"))
async def start_cmd(message: types.Message):
    user_data[message.from_user.id] = []
    await message.answer("Привет! Выберите действие:", reply_markup=main_menu())

# Обработка кнопок
@dp.callback_query()
async def menu_handler(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id

    if callback.data == "create_excel":
        user_data[user_id] = []
        await callback.message.answer("Создание Excel. Добавьте товар:", reply_markup=product_menu())

    elif callback.data == "add_item":
        await callback.message.answer("Отправьте фото или файл товара:")
        await state.set_state(ProductForm.waiting_for_photo)

    elif callback.data == "edit_item":
        if not user_data[user_id]:
            await callback.message.answer("Список товаров пуст!")
            return
        kb = InlineKeyboardMarkup()
        for i, item in enumerate(user_data[user_id], start=1):
            kb.add(InlineKeyboardButton(text=f"{i}. {item['color']} | {item['size']}", callback_data=f"edit_{i-1}"))
        await callback.message.answer("Выберите товар для редактирования:", reply_markup=kb)
        await state.set_state(EditProduct.choosing_product)

    elif callback.data == "delete_item":
        if not user_data[user_id]:
            await callback.message.answer("Список товаров пуст!")
            return
        kb = InlineKeyboardMarkup()
        for i, item in enumerate(user_data[user_id], start=1):
            kb.add(InlineKeyboardButton(text=f"{i}. {item['color']} | {item['size']}", callback_data=f"del_{i-1}"))
        await callback.message.answer("Выберите товар для удаления:", reply_markup=kb)
        await state.set_state(DeleteProduct.choosing_product)

    elif callback.data == "finish_excel":
        if not user_data[user_id]:
            await callback.message.answer("Список товаров пуст!")
            return
        text = "📦 Товары:\n"
        for i, item in enumerate(user_data[user_id], start=1):
            text += f"{i}. {item['color']} | {item['size']} | {item['qty']} шт\n"
        await callback.message.answer(text, reply_markup=finish_menu())

    elif callback.data == "back_to_products":
        await callback.message.answer("Меню работы с товарами:", reply_markup=product_menu())

    elif callback.data == "save_excel":
        await callback.message.answer("Введите код груза:")
        await state.set_state(SaveExcel.waiting_for_code)

    elif callback.data == "view_archive":
        user_folder = f"archive/{user_id}"
        if not os.path.exists(user_folder) or not os.listdir(user_folder):
            await callback.message.answer("Ваш архив пуст 📂")
        else:
            files = os.listdir(user_folder)
            text = "📜 Ваши файлы:\n" + "\n".join(files)
            await callback.message.answer(text)

# ----------------- ДОБАВЛЕНИЕ -----------------
@dp.message(ProductForm.waiting_for_photo)
async def process_photo(message: types.Message, state: FSMContext):
    if message.photo:
        file_id = message.photo[-1].file_id
    elif message.document:
        file_id = message.document.file_id
    else:
        await message.answer("Пожалуйста, отправьте фото или файл 📷")
        return

    file = await bot.get_file(file_id)
    downloaded = await bot.download_file(file.file_path)
    await state.update_data(photo=downloaded.read())

    await message.answer("Введите ссылку на товар:")
    await state.set_state(ProductForm.waiting_for_link)

@dp.message(ProductForm.waiting_for_link)
async def process_link(message: types.Message, state: FSMContext):
    await state.update_data(link=message.text.strip())
    await message.answer("Введите цвет товара:")
    await state.set_state(ProductForm.waiting_for_color)

@dp.message(ProductForm.waiting_for_color)
async def process_color(message: types.Message, state: FSMContext):
    await state.update_data(color=message.text.strip())
    await message.answer("Введите размер товара:")
    await state.set_state(ProductForm.waiting_for_size)

@dp.message(ProductForm.waiting_for_size)
async def process_size(message: types.Message, state: FSMContext):
    await state.update_data(size=message.text.strip())
    await message.answer("Введите количество товара:")
    await state.set_state(ProductForm.waiting_for_qty)

@dp.message(ProductForm.waiting_for_qty)
async def process_qty(message: types.Message, state: FSMContext):
    try:
        qty = int(message.text.strip())
    except ValueError:
        await message.answer("Введите число!")
        return
    await state.update_data(qty=qty)
    await message.answer("Введите комментарий:")
    await state.set_state(ProductForm.waiting_for_comment)

@dp.message(ProductForm.waiting_for_comment)
async def process_comment(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    await state.update_data(comment=message.text.strip())

    data = await state.get_data()
    product = {
        "photo": data["photo"],
        "link": data["link"],
        "color": data["color"],
        "size": data["size"],
        "qty": data["qty"],
        "comment": data["comment"]
    }

    user_data[user_id].append(product)

    await state.clear()
    await message.answer("✅ Товар добавлен!", reply_markup=product_menu())

# ----------------- РЕДАКТИРОВАНИЕ -----------------
@dp.callback_query(EditProduct.choosing_product)
async def choose_product_to_edit(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    await state.update_data(edit_index=idx)

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Фото", callback_data="field_photo")],
        [InlineKeyboardButton(text="Ссылка", callback_data="field_link")],
        [InlineKeyboardButton(text="Цвет", callback_data="field_color")],
        [InlineKeyboardButton(text="Размер", callback_data="field_size")],
        [InlineKeyboardButton(text="Количество", callback_data="field_qty")],
        [InlineKeyboardButton(text="Комментарий", callback_data="field_comment")]
    ])
    await callback.message.answer("Что хотите изменить?", reply_markup=kb)
    await state.set_state(EditProduct.choosing_field)

@dp.callback_query(EditProduct.choosing_field)
async def choose_field(callback: types.CallbackQuery, state: FSMContext):
    field = callback.data.split("_")[1]
    await state.update_data(field=field)
    if field == "photo":
        await callback.message.answer("Отправьте новое фото/файл:")
    else:
        await callback.message.answer("Введите новое значение:")
    await state.set_state(EditProduct.updating_value)

@dp.message(EditProduct.updating_value)
async def update_value(message: types.Message, state: FSMContext):
    data = await state.get_data()
    user_id = message.from_user.id
    idx = data["edit_index"]
    field = data["field"]

    if field == "photo":
        if message.photo:
            file_id = message.photo[-1].file_id
        elif message.document:
            file_id = message.document.file_id
        else:
            await message.answer("Пожалуйста, отправьте фото или файл 📷")
            return
        file = await bot.get_file(file_id)
        downloaded = await bot.download_file(file.file_path)
        user_data[user_id][idx]["photo"] = downloaded.read()
    elif field == "qty":
        try:
            user_data[user_id][idx]["qty"] = int(message.text.strip())
        except ValueError:
            await message.answer("Введите число!")
            return
    else:
        user_data[user_id][idx][field] = message.text.strip()

    await state.clear()
    await message.answer("✅ Товар обновлён!", reply_markup=product_menu())

# ----------------- УДАЛЕНИЕ -----------------
@dp.callback_query(DeleteProduct.choosing_product)
async def delete_product(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    user_id = callback.from_user.id
    deleted = user_data[user_id].pop(idx)
    await state.clear()
    await callback.message.answer(f"🗑 Товар удалён: {deleted['color']} | {deleted['size']}", reply_markup=product_menu())

# ----------------- СОХРАНЕНИЕ EXCEL -----------------
@dp.message(SaveExcel.waiting_for_code)
async def save_excel_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    code = message.text.strip()
    filename = f"{code}.xlsx"

    wb = Workbook()
    ws = wb.active

    # Шапка компании
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    cell1 = ws["A1"]
    cell1.value = "MAGEZ"
    cell1.font = Font(color="FFFFFF", bold=True, size=16)
    cell1.fill = PatternFill("solid", fgColor="FF0000")
    cell1.alignment = Alignment(horizontal="center", vertical="center")

    cell2 = ws["A2"]
    cell2.value = "Торгово-Логистическая компания"
    cell2.font = Font(color="FFFFFF", bold=True, size=12)
    cell2.fill = PatternFill("solid", fgColor="FF0000")
    cell2.alignment = Alignment(horizontal="center", vertical="center")

    # Стили
    header_fill = PatternFill("solid", fgColor="808080")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Заголовки
    headers = ["Фото", "Ссылка", "Цвет", "Размер", "Количество", "Комментарий"]
    ws.append([])
    ws.append(headers)
    for col in range(1, 7):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Товары
    for i, item in enumerate(user_data[user_id], start=5):
        ws.cell(row=i, column=2).value = item["link"]
        ws.cell(row=i, column=3).value = item["color"]
        ws.cell(row=i, column=4).value = item["size"]
        ws.cell(row=i, column=5).value = item["qty"]
        ws.cell(row=i, column=6).value = item["comment"]

        for col in range(2, 7):
            cell = ws.cell(row=i, column=col)
            cell.alignment = center_align
            cell.border = thin_border

        if item["photo"]:
            img = XLImage(BytesIO(item["photo"]))
            img.width, img.height = 80, 80
            ws.add_image(img, f"A{i}")
            ws.row_dimensions[i].height = 60

    # Автоширина
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    # Архив
    user_folder = f"archive/{user_id}"
    os.makedirs(user_folder, exist_ok=True)
    filepath = f"{user_folder}/{filename}"
    wb.save(filepath)

    await message.answer_document(InputFile(filepath))
    await message.answer("✅ Файл сохранён в архив!", reply_markup=main_menu())
    await state.clear()

# 🚀 Запуск
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
